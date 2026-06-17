"""Write a nicely formatted .xlsx: bold/centered/colored header, clickable links,
auto column widths, frozen header, autofilter, and color-coded Exp Fit cells.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("Fit%", "score", 7),
    ("Posted", "posted_date", 13),
    ("Title", "title", 46),
    ("Company / Location", "location", 38),
    ("Req Exp", "req_exp", 14),
    ("Exp Fit", "exp_fit", 16),
    ("Source", "source", 18),
    ("Link", "url", 60),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")      # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LINK_FONT = Font(color="0563C1", underline="single")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FIT_FILL = {
    "Fits": PatternFill("solid", fgColor="C6EFCE"),          # green
    "Over-qualified": PatternFill("solid", fgColor="FFF2CC"),  # pale yellow
    "Unknown": PatternFill("solid", fgColor="F2F2F2"),        # grey
}
STRETCH_FILL = PatternFill("solid", fgColor="FCE4D6")        # orange-ish


def write_xlsx(rows, path, title="PM Roles"):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # header
    for c, (label, _, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 26

    # body
    for r, row in enumerate(rows, start=2):
        for c, (_, key, _) in enumerate(COLUMNS, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            if key == "url":
                cell.value = val
                cell.hyperlink = val
                cell.font = LINK_FONT
                cell.alignment = LEFT
            elif key == "score":
                cell.value = int(val) if str(val).isdigit() else val
                cell.alignment = CENTER
            elif key in ("posted_date", "req_exp", "exp_fit"):
                cell.value = val
                cell.alignment = CENTER
                if key == "exp_fit":
                    if str(val).startswith("Stretch"):
                        cell.fill = STRETCH_FILL
                    elif val in FIT_FILL:
                        cell.fill = FIT_FILL[val]
            else:
                cell.value = val
                cell.alignment = LEFT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    wb.save(path)
    return path
