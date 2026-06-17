"""Filter the latest scan Excel to genuine India PM roles and export a focused
India Excel (Fits first, latest date, clickable links)."""
import glob, os, re

import openpyxl
import xlsx_export

# newest scan .xlsx, but not a previously-generated INDIA_ file
cands = [p for p in glob.glob("results/*.xlsx") if not os.path.basename(p).startswith("INDIA")]
latest = max(cands, key=os.path.getmtime)
wb = openpyxl.load_workbook(latest)
ws = wb.active
headers = [c.value for c in ws[1]]
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, r))
    rows.append({"score": d["Fit%"], "posted_date": d["Posted"], "title": d["Title"],
                 "location": d["Company / Location"], "req_exp": d["Req Exp"],
                 "exp_fit": d["Exp Fit"], "source": d["Source"], "url": d["Link"]})

CITY = re.compile(r"india|bengaluru|bangalore|gurugram|gurgaon|mumbai|hyderabad|pune|delhi|noida|"
                  r"chennai|kolkata|jaipur|kochi|ahmedabad", re.I)
FOREIGN = re.compile(r"canada|united states|\bus\b|\buk\b|europe|emea|singapore|dublin|london|"
                     r"germany|poland|spain|argentina|cairo|malta|athens|warsaw", re.I)


def is_india(r):
    loc = r["location"] or ""
    if FOREIGN.search(loc) and not CITY.search(loc):
        return False
    return bool(CITY.search(loc)) or (r["source"] == "linkedin" and "remote" in loc.lower())


ind = [r for r in rows if is_india(r)]
ind.sort(key=lambda r: (1 if r["exp_fit"] == "Fits" else 0,
                        str(r["posted_date"]) if r["posted_date"] != "unknown" else "0000",
                        int(r["score"])), reverse=True)

out = "results/INDIA_PM_roles.xlsx"
xlsx_export.write_xlsx(ind, out, title="India PM Roles")
fits = sum(1 for r in ind if r["exp_fit"] == "Fits")
print(f"Read: {latest} ({len(rows)} rows) | India-only: {len(ind)} | Fits: {fits}")
print(f"Saved Excel: {os.path.abspath(out)}")
